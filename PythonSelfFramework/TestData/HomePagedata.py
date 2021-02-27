import openpyxl


class HomePageData:

   # test_HomePage_data = [{"firstname":"Smita","lastname":"Jadhav", "gender":"Female"}, {"firstname":"Vivaan","lastname":"sharma", "gender":"male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\pythonProject\\SMSheet1.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        #print(Dict)
        return[Dict]

